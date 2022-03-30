import os
import pandas as pd
import pathlib
import re
from openpyxl import Workbook, load_workbook

#global variables
SPEC_PATH = input("path: ")


def format_date(extracted_string):
    new_s = re.sub(r"[^a-zA-Z0-9 ]", "", extracted_string)
    return new_s


def extract_date(workbook_name):
    for files in os.listdir(SPEC_PATH):
        if files.endswith(".xls") or x.endswith(".xlsx"):
            wb = load_workbook(files)
            ws = wb.active
            return format_date(ws["C5"].value)


# main function
SPEC_PATH = pathlib.Path(SPEC_PATH).resolve()  # will access the given path
count = 0

user_choice = input(
    "Would you like the new name to contain a string before the date?(y/n) ").lower()

while True:
    if user_choice.isdigit() or (user_choice != 'y' or user_choice != 'n'):
        print("Please enter a valid number!")
        if user_choice == 'y':
            head_string = input("Please enter the string: ")

            for x in os.listdir(SPEC_PATH):
                # might include all xl extensions just to be safe
                if x.endswith(".xls") or x.endswith(".xlsx"):
                    extract_date(x)
            for file in SPEC_PATH:
                if file.is_file():
                    custom_name = extract_date(file)
                    new_name = f"{head_string}{file.suffix}"
                    if SPEC_PATH.is_file():
                        continue
                    else:
                        try:
                            file.rename(SPEC_PATH/new_name)
                            print("Success!")
                        except OSError as e:
                            print(e)

        elif user_choice == 'n':
            # will count all the files that contain a specified extension
            for x in os.listdir(SPEC_PATH):
                # might include all xl extensions just to be safe
                if x.endswith(".xls") or x.endswith(".xlsx"):
                    count += 1
                    extract_date(x)
            for file in SPEC_PATH:
                if file.is_file():
                    custom_name = extract_date(file)
                    new_name = f"{file.suffix}"
                    if SPEC_PATH.is_file():
                        continue
                    else:
                        try:
                            file.rename(SPEC_PATH/new_name)
                            print("Success!")
                        except OSError as e:
                            print(e)
            print(count)
        else:
            pass
    else:
        break
